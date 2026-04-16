"""
Fiets-admin router: Excel-import en export van reserveringsdata.
- Import: altijd mogelijk, duplicaten worden overgeslagen (zelfde gast + fiets + datums)
- Export: alle reserveringen als xlsx
"""
import asyncio
import io
import re
from datetime import date, datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from ..database import get_db
from ..models import (
    Bike, BikeType, BikeReservation, BikeReservationBike,
    BikeReservationStatus, SystemSetting,
)
from ..auth import RequireUser

router = APIRouter(prefix="/bike-admin", tags=["bike-admin"])

# Dutch month names → month number
MONTH_MAP = {
    "JANUARI": 1, "FEBRUARI": 2, "MAART": 3, "APRIL": 4,
    "MEI": 5, "JUNI": 6, "JULI": 7, "AUGUSTUS": 8,
    "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}

SKIP_VALUES = {"onderhoud", "x", "reserve", ""}


def _extract_bike_number(header_text: str) -> str | None:
    if not header_text:
        return None
    m = re.search(r"fiets\s+(\d+)", str(header_text).lower())
    return m.group(1) if m else None


def _normalize_name(name) -> str | None:
    if name is None:
        return None
    s = str(name).strip()
    if not s or s.lower() in SKIP_VALUES:
        return None
    return s


def _parse_sections(ws) -> list[dict]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    sections = []
    section_starts = []
    for i, row in enumerate(rows):
        if row and len(row) > 1 and isinstance(row[1], datetime):
            section_starts.append(i)

    for sec_idx, start in enumerate(section_starts):
        dt: datetime = rows[start][1]
        year, month = dt.year, dt.month

        header_row_idx = start + 2
        if header_row_idx >= len(rows):
            continue
        header_row = rows[header_row_idx]

        col_map: dict[int, str] = {}
        if header_row:
            for col_i, cell in enumerate(header_row):
                num = _extract_bike_number(cell)
                if num:
                    col_map[col_i] = num

        if not col_map:
            continue

        data_start = header_row_idx + 1
        data_end = section_starts[sec_idx + 1] if sec_idx + 1 < len(section_starts) else len(rows)

        data_rows = []
        for row_i in range(data_start, data_end):
            row = rows[row_i]
            if row and len(row) > 1 and isinstance(row[1], int):
                data_rows.append((row[1], row))

        sections.append({
            "year": year,
            "month": month,
            "col_map": col_map,
            "data_rows": data_rows,
        })
    return sections


def _extract_segments(data_rows: list, col_idx: int) -> list[dict]:
    segments = []
    current = None
    for day, row in data_rows:
        name = _normalize_name(row[col_idx]) if col_idx < len(row) else None
        if name and current and name.lower() == current["guest_name"].lower():
            current["end_day"] = day
        elif name:
            if current:
                segments.append(current)
            current = {"start_day": day, "end_day": day, "guest_name": name}
        else:
            if current:
                segments.append(current)
            current = None
    if current:
        segments.append(current)
    return segments


def _parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse Excel bytes → lijst van reservation-dicts + foutmeldingen. Synchrone functie."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    results = []
    errors = []
    today = date.today()

    for sheet_name in wb.sheetnames:
        if sheet_name == "Blad2":
            continue
        ws = wb[sheet_name]
        sections = _parse_sections(ws)

        for section in sections:
            year = section["year"]
            month = section["month"]
            col_map = section["col_map"]
            data_rows = section["data_rows"]

            for col_idx, bike_number in col_map.items():
                segments = _extract_segments(data_rows, col_idx)
                for seg in segments:
                    try:
                        start_date = date(year, month, int(seg["start_day"]))
                        end_date = date(year, month, int(seg["end_day"]))
                    except (ValueError, OverflowError) as e:
                        errors.append(f"Ongeldige datum in {sheet_name}: {e}")
                        continue
                    num_days = (end_date - start_date).days + 1
                    status = (
                        BikeReservationStatus.completed
                        if end_date < today
                        else BikeReservationStatus.active
                    )
                    results.append({
                        "bike_number": bike_number,
                        "guest_name": seg["guest_name"],
                        "start_date": start_date,
                        "end_date": end_date,
                        "num_days": num_days,
                        "status": status,
                    })
    return results, errors


def _build_export_excel(reservations: list[dict]) -> bytes:
    """Genereer xlsx-bestand van alle reserveringen. Synchrone functie."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reserveringen"

    headers = [
        "ID", "Gast", "Kamer", "Startdatum", "Einddatum",
        "Dagen", "# Fietsen", "Fietstype", "Status", "Fietsnummers", "Notities"
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_nl = {
        "active": "Actief",
        "completed": "Afgerond",
        "cancelled": "Geannuleerd",
    }

    for row_i, res in enumerate(reservations, 2):
        ws.cell(row=row_i, column=1, value=res["id"])
        ws.cell(row=row_i, column=2, value=res["guest_name"])
        ws.cell(row=row_i, column=3, value=res["guest_room"] or "")
        ws.cell(row=row_i, column=4, value=res["start_date"])
        ws.cell(row=row_i, column=5, value=res["end_date"])
        ws.cell(row=row_i, column=6, value=res["num_days"])
        ws.cell(row=row_i, column=7, value=res["num_bikes"])
        ws.cell(row=row_i, column=8, value=res["bike_type_name"] or "")
        ws.cell(row=row_i, column=9, value=status_nl.get(res["status"], res["status"]))
        ws.cell(row=row_i, column=10, value=res["bike_numbers"] or "")
        ws.cell(row=row_i, column=11, value=res["notes"] or "")

        # Datum-cellen opmaken
        for col in (4, 5):
            ws.cell(row=row_i, column=col).number_format = "DD-MM-YYYY"

        # Rij-kleur op basis van status
        if res["status"] == "completed":
            fill = PatternFill("solid", fgColor="F0F9F0")
        elif res["status"] == "cancelled":
            fill = PatternFill("solid", fgColor="FFF0F0")
        else:
            fill = PatternFill("solid", fgColor="F0F4FF")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_i, column=col).fill = fill

    # Kolombreedtes
    col_widths = [8, 25, 10, 14, 14, 8, 10, 18, 14, 20, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.post("/import-excel")
async def import_excel(
    user: RequireUser,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not user.is_admin:
        raise HTTPException(403, "Alleen admins kunnen Excel importeren")
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Alleen .xlsx bestanden zijn toegestaan")

    contents = await file.read()

    # Parse synchroon in thread (openpyxl is niet async)
    parsed, errors = await asyncio.to_thread(_parse_excel, contents)

    # Bouw bike_number → Bike lookup
    bikes_result = await db.execute(select(Bike))
    bike_lookup: dict[str, Bike] = {b.number: b for b in bikes_result.scalars().all()}

    # Bouw deduplicatie-set: (bike_id, start_date, end_date)
    existing_result = await db.execute(
        select(BikeReservationBike.bike_id, BikeReservation.start_date, BikeReservation.end_date)
        .join(BikeReservation, BikeReservation.id == BikeReservationBike.reservation_id)
    )
    existing_keys: set[tuple] = {
        (row.bike_id, row.start_date, row.end_date)
        for row in existing_result
    }

    # Haal huidige max ID op voor counter
    max_id_result = await db.execute(select(func.max(BikeReservation.id)))
    last_id = max_id_result.scalar() or 0
    next_id = last_id + 1

    imported = 0
    skipped_dup = 0
    skipped_no_bike = 0

    for entry in parsed:
        bike = bike_lookup.get(entry["bike_number"])
        if not bike:
            skipped_no_bike += 1
            continue

        dedup_key = (bike.id, entry["start_date"], entry["end_date"])
        if dedup_key in existing_keys:
            skipped_dup += 1
            continue

        res = BikeReservation(
            id=next_id,
            guest_name=entry["guest_name"],
            guest_room=None,
            start_date=entry["start_date"],
            end_date=entry["end_date"],
            num_days=entry["num_days"],
            num_bikes=1,
            bike_type_id=bike.type_id,
            status=entry["status"],
            notes="Geïmporteerd uit Excel",
        )
        db.add(res)
        await db.flush()
        db.add(BikeReservationBike(reservation_id=res.id, bike_id=bike.id))
        bike.total_rental_days += entry["num_days"]
        existing_keys.add(dedup_key)
        next_id += 1
        imported += 1

    # Update counter instelling
    counter = await db.get(SystemSetting, "bike_reservation_counter")
    if counter:
        counter.value = str(next_id)
    else:
        db.add(SystemSetting(key="bike_reservation_counter", value=str(next_id)))

    await db.commit()

    return {
        "ok": True,
        "imported": imported,
        "skipped": skipped_dup + skipped_no_bike,
        "skipped_duplicates": skipped_dup,
        "skipped_no_bike": skipped_no_bike,
        "errors": errors[:10],
    }


@router.get("/export-excel")
async def export_excel(
    user: RequireUser,
    db: AsyncSession = Depends(get_db),
):
    """Exporteer alle reserveringen als xlsx-bestand."""
    if not user.is_admin:
        raise HTTPException(403, "Alleen admins kunnen exporteren")

    # Haal alle reserveringen op met fietstype + fietsnummers
    res_result = await db.execute(
        select(BikeReservation).order_by(BikeReservation.start_date.desc())
    )
    reservations = res_result.scalars().all()

    # Fietstype lookup
    types_result = await db.execute(select(BikeType))
    type_lookup: dict[int, str] = {t.id: t.name for t in types_result.scalars().all()}

    # Bike lookup
    bikes_res = await db.execute(select(Bike))
    bike_lookup_id: dict[int, str] = {b.id: b.number for b in bikes_res.scalars().all()}

    # Koppeltabel: reservation_id → list[bike_id]
    links_result = await db.execute(select(BikeReservationBike))
    bike_links: dict[int, list[int]] = {}
    for link in links_result.scalars().all():
        bike_links.setdefault(link.reservation_id, []).append(link.bike_id)

    rows = []
    for r in reservations:
        linked_bikes = bike_links.get(r.id, [])
        bike_numbers = ", ".join(
            f"#{bike_lookup_id[bid]}" for bid in linked_bikes if bid in bike_lookup_id
        )
        rows.append({
            "id": r.id,
            "guest_name": r.guest_name,
            "guest_room": r.guest_room,
            "start_date": r.start_date,
            "end_date": r.end_date,
            "num_days": r.num_days,
            "num_bikes": r.num_bikes,
            "bike_type_name": type_lookup.get(r.bike_type_id) if r.bike_type_id else None,
            "status": r.status.value if hasattr(r.status, "value") else r.status,
            "bike_numbers": bike_numbers,
            "notes": r.notes,
        })

    excel_bytes = await asyncio.to_thread(_build_export_excel, rows)

    today_str = date.today().strftime("%Y-%m-%d")
    filename = f"fietsreserveringen_{today_str}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
