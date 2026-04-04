import io
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
import openpyxl

from ..database import get_db
from ..models import Ticket, Status, Category, Priority
from ..auth import RequireUser

router = APIRouter(prefix="/reports", tags=["reports"])


def _date_filter(from_date: date | None, to_date: date | None):
    filters = []
    if from_date:
        filters.append(Ticket.created_at >= datetime.combine(from_date, datetime.min.time()))
    if to_date:
        filters.append(Ticket.created_at <= datetime.combine(to_date, datetime.max.time()))
    return filters


@router.get("/summary")
async def get_summary(
    user: RequireUser,
    db: AsyncSession = Depends(get_db),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
):
    """Overzichtsstatistieken voor het dashboard."""
    date_filters = _date_filter(from_date, to_date)

    # Totalen per status
    status_counts = {}
    for s in Status:
        count = await db.scalar(select(func.count()).where(and_(Ticket.status == s, *date_filters)))
        status_counts[s.value] = count or 0

    # Totalen per categorie
    category_counts = {}
    for c in Category:
        count = await db.scalar(select(func.count()).where(and_(Ticket.category == c, *date_filters)))
        category_counts[c.value] = count or 0

    # Totalen per prioriteit
    priority_counts = {}
    for p in Priority:
        count = await db.scalar(select(func.count()).where(and_(Ticket.priority == p, *date_filters)))
        priority_counts[p.value] = count or 0

    # Gemiddelde afdoeningstijd (in uren) voor gesloten tickets
    closed_tickets = (
        await db.execute(
            select(Ticket.created_at, Ticket.closed_at).where(
                and_(Ticket.status == Status.closed, Ticket.closed_at.isnot(None), *date_filters)
            )
        )
    ).all()

    avg_resolution_hours = None
    if closed_tickets:
        durations = [
            (row.closed_at - row.created_at).total_seconds() / 3600
            for row in closed_tickets
            if row.closed_at and row.created_at
        ]
        avg_resolution_hours = round(sum(durations) / len(durations), 1) if durations else None

    return {
        "status_counts": status_counts,
        "category_counts": category_counts,
        "priority_counts": priority_counts,
        "avg_resolution_hours": avg_resolution_hours,
        "total_tickets": sum(status_counts.values()),
    }


@router.get("/timeline")
async def get_timeline(
    user: RequireUser,
    db: AsyncSession = Depends(get_db),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
    group_by: str = Query("day", regex="^(day|week|month)$"),
):
    """Tickets per tijdseenheid voor de lijngrafiek."""
    date_filters = _date_filter(from_date, to_date)

    result = await db.execute(
        select(Ticket.created_at, Ticket.status, Ticket.category).where(and_(*date_filters))
    )
    rows = result.all()

    # Groepeer in Python (SQLite heeft beperkte date_trunc ondersteuning)
    timeline: dict[str, dict] = {}
    for row in rows:
        if group_by == "day":
            key = row.created_at.strftime("%Y-%m-%d")
        elif group_by == "week":
            key = row.created_at.strftime("%Y-W%W")
        else:
            key = row.created_at.strftime("%Y-%m")

        if key not in timeline:
            timeline[key] = {"period": key, "total": 0, "open": 0, "closed": 0}
        timeline[key]["total"] += 1
        if row.status == Status.closed:
            timeline[key]["closed"] += 1
        else:
            timeline[key]["open"] += 1

    return sorted(timeline.values(), key=lambda x: x["period"])


@router.get("/export/csv")
async def export_csv(
    user: RequireUser,
    db: AsyncSession = Depends(get_db),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
):
    """Exporteer tickets als CSV."""
    date_filters = _date_filter(from_date, to_date)
    result = await db.execute(
        select(Ticket).where(and_(*date_filters)).order_by(Ticket.created_at.desc())
    )
    tickets = result.scalars().all()

    lines = ["id,titel,categorie,status,prioriteit,locatie,aangemaakt_door,toegewezen_aan,aangemaakt_op,gesloten_op"]
    for t in tickets:
        lines.append(",".join([
            t.id, f'"{t.title}"', t.category.value, t.status.value, t.priority.value,
            t.location_id or "", t.created_by, t.assigned_to or "",
            t.created_at.isoformat(), t.closed_at.isoformat() if t.closed_at else "",
        ]))

    content = "\n".join(lines).encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=tickets.csv"},
    )


@router.get("/export/excel")
async def export_excel(
    user: RequireUser,
    db: AsyncSession = Depends(get_db),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
):
    """Exporteer tickets als Excel bestand."""
    date_filters = _date_filter(from_date, to_date)
    result = await db.execute(
        select(Ticket).where(and_(*date_filters)).order_by(Ticket.created_at.desc())
    )
    tickets = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = ["ID", "Titel", "Beschrijving", "Categorie", "Status", "Prioriteit",
               "Locatie", "Aangemaakt door", "Toegewezen aan", "Aangemaakt op", "Gesloten op"]
    ws.append(headers)

    # Opmaak header
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    for t in tickets:
        ws.append([
            t.id, t.title, t.description or "", t.category.value, t.status.value,
            t.priority.value, t.location_id or "", t.created_by, t.assigned_to or "",
            t.created_at.isoformat(), t.closed_at.isoformat() if t.closed_at else "",
        ])

    # Auto kolombreedte
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tickets.xlsx"},
    )
